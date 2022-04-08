import openpyxl

# Create an empty workbook
from ope

workbook = openpyxl.Workbook()

# Get the current sheet for this workbook
sheet = workbook.active
sheet.title = "first_sheet"
for i, columname in enumerate(('Nr', 'Name')):
    sheet.cell(1, i + 1, columname)

for i, user_name in enumerate(('a', 'b')):
    y_offset = 2 + i
    sheet.cell(y_offset, 1, i)
    sheet.cell(y_offset, 2, user_name)
second_sheet = workbook.create_sheet('second_sheet')

workbook.save("demo.xlsx")

from open

workbook2 = openpyxl.open("demo.xlsx")
